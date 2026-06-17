import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# =========================
# Paths
# =========================
PHASE_XLSX = "/data/drdcad/Hyuna/projects/vessel_seg/data/contrast_phase_check/phase_summary.xlsx"
TARGET_XLSX = "/data/drdcad/Hyuna/projects/vessel_seg/data/5samples_updated.xlsx"
TARGET_SHEET = "gt_manual"

OUT_XLSX = "/data/drdcad/Hyuna/projects/vessel_seg/data/5samples_updated_with_phase.xlsx"


# =========================
# Helpers
# =========================
def norm_col(c):
    return str(c).strip().lower().replace(" ", "_")


def find_col(columns, candidates):
    """
    Find column by flexible matching.
    """
    norm_map = {norm_col(c): c for c in columns}

    for cand in candidates:
        cand_norm = norm_col(cand)

        # exact
        if cand_norm in norm_map:
            return norm_map[cand_norm]

        # contains
        for k, orig in norm_map.items():
            if cand_norm in k:
                return orig

    raise ValueError(f"Cannot find column among candidates: {candidates}\nExisting columns: {list(columns)}")


def normalize_id(v, width=None):
    """
    Excel에서 1, 1.0, '00001' 모두 안정적으로 처리.
    """
    if pd.isna(v):
        return None

    s = str(v).strip()

    # 1.0 같은 경우 보정
    try:
        if s.replace(".", "", 1).isdigit():
            s = str(int(float(s)))
    except Exception:
        pass

    if width is not None:
        s = s.zfill(width)

    return s


def make_key(subject, study, series):
    subject = normalize_id(subject, width=5)
    study = normalize_id(study, width=5)
    series = normalize_id(series, width=None)

    if subject is None or study is None or series is None:
        return None

    return f"{subject}_{study}_{series}"


# =========================
# 1) Read phase_summary
# =========================
phase_df = pd.read_excel(PHASE_XLSX, dtype=str)

print("phase_summary columns:")
print(list(phase_df.columns))

phase_subject_col = find_col(phase_df.columns, ["subject_id", "subject"])
phase_study_col = find_col(phase_df.columns, ["study_id", "study"])
phase_series_col = find_col(phase_df.columns, ["series_id", "series"])
phase_col = find_col(phase_df.columns, ["phase"])

phase_map = {}

for _, row in phase_df.iterrows():
    key = make_key(
        row[phase_subject_col],
        row[phase_study_col],
        row[phase_series_col],
    )

    if key is None:
        continue

    phase_val = row[phase_col]

    if pd.isna(phase_val):
        continue

    phase_map[key] = str(phase_val).strip()

print(f"Loaded phase values: {len(phase_map)}")


# =========================
# 2) Open target workbook and gt_manual sheet
# =========================
wb = load_workbook(TARGET_XLSX)

if TARGET_SHEET not in wb.sheetnames:
    raise ValueError(f"Sheet '{TARGET_SHEET}' not found. Existing sheets: {wb.sheetnames}")

ws = wb[TARGET_SHEET]

# Header row = first row
headers = [cell.value for cell in ws[1]]
print("gt_manual columns:")
print(headers)

target_subject_col_name = find_col(
    headers,
    ["anon_patient_ID", "subject_id", "subject", "patient"]
)

target_study_col_name = find_col(
    headers,
    ["anon_study_ID", "study_id", "study"]
)

target_series_col_name = find_col(
    headers,
    ["SeriesNumber", "series_id", "series"]
)

# openpyxl column index는 1-based
header_to_idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

subject_idx = header_to_idx[str(target_subject_col_name).strip()]
study_idx = header_to_idx[str(target_study_col_name).strip()]
series_idx = header_to_idx[str(target_series_col_name).strip()]

# phase column 있으면 업데이트, 없으면 새로 생성
existing_phase_cols = [
    i + 1 for i, h in enumerate(headers)
    if h is not None and norm_col(h) == "phase"
]

if existing_phase_cols:
    phase_idx = existing_phase_cols[0]
    print(f"Using existing phase column at index {phase_idx}")
else:
    phase_idx = ws.max_column + 1
    ws.cell(row=1, column=phase_idx).value = "phase"
    print(f"Created new phase column at index {phase_idx}")


# =========================
# 3) Match and copy phase
# =========================
matched = 0
missing = 0
missing_keys = []

for r in range(2, ws.max_row + 1):
    subject = ws.cell(row=r, column=subject_idx).value
    study = ws.cell(row=r, column=study_idx).value
    series = ws.cell(row=r, column=series_idx).value

    key = make_key(subject, study, series)

    if key is None:
        continue

    if key in phase_map:
        ws.cell(row=r, column=phase_idx).value = phase_map[key]
        matched += 1
    else:
        missing += 1
        missing_keys.append(key)

print(f"Matched rows: {matched}")
print(f"Rows without phase match: {missing}")

if missing_keys:
    print("First 20 missing keys:")
    for k in missing_keys[:20]:
        print(k)

# =========================
# 3-2) Copy non-native phase from gt_manual to used_data
# =========================

USED_SHEET = "used_data"
CLEAR_USED_DATA_PHASE = True  # True면 used_data의 기존 phase 값을 먼저 비움

if USED_SHEET not in wb.sheetnames:
    raise ValueError(f"Sheet '{USED_SHEET}' not found. Existing sheets: {wb.sheetnames}")

ws_gt = wb[TARGET_SHEET]
ws_used = wb[USED_SHEET]

# -------------------------
# Helper for openpyxl sheets
# -------------------------
def get_headers(ws):
    return [cell.value for cell in ws[1]]

def get_col_idx(headers, candidates):
    col_name = find_col(headers, candidates)
    header_to_idx = {
        str(h).strip(): i + 1
        for i, h in enumerate(headers)
        if h is not None
    }
    return header_to_idx[str(col_name).strip()], col_name

def get_or_create_col(ws, col_name):
    headers = get_headers(ws)

    for i, h in enumerate(headers):
        if h is not None and norm_col(h) == norm_col(col_name):
            return i + 1

    new_idx = ws.max_column + 1
    ws.cell(row=1, column=new_idx).value = col_name
    return new_idx


# -------------------------
# gt_manual columns
# -------------------------
gt_headers = get_headers(ws_gt)

gt_subject_idx, _ = get_col_idx(
    gt_headers,
    ["anon_patient_ID", "subject_id", "subject", "patient"]
)
gt_study_idx, _ = get_col_idx(
    gt_headers,
    ["anon_study_ID", "study_id", "study"]
)
gt_series_idx, _ = get_col_idx(
    gt_headers,
    ["SeriesNumber", "series_id", "series"]
)
gt_phase_idx, _ = get_col_idx(
    gt_headers,
    ["phase"]
)

# -------------------------
# Build non-native phase map from gt_manual
# -------------------------
non_native_phase_map = {}

for r in range(2, ws_gt.max_row + 1):
    subject = ws_gt.cell(row=r, column=gt_subject_idx).value
    study = ws_gt.cell(row=r, column=gt_study_idx).value
    series = ws_gt.cell(row=r, column=gt_series_idx).value
    phase_val = ws_gt.cell(row=r, column=gt_phase_idx).value

    key = make_key(subject, study, series)

    if key is None or phase_val is None:
        continue

    phase_str = str(phase_val).strip()

    if phase_str == "":
        continue

    # native는 used_data로 복사하지 않음
    if phase_str.lower() == "native":
        continue

    non_native_phase_map[key] = phase_str

print(f"Non-native phase values from gt_manual: {len(non_native_phase_map)}")


# -------------------------
# used_data columns
# -------------------------
used_headers = get_headers(ws_used)

used_subject_idx, _ = get_col_idx(
    used_headers,
    ["anon_patient_ID", "subject_id", "subject", "patient"]
)
used_study_idx, _ = get_col_idx(
    used_headers,
    ["anon_study_ID", "study_id", "study"]
)
used_series_idx, _ = get_col_idx(
    used_headers,
    ["SeriesNumber", "series_id", "series"]
)

used_phase_idx = get_or_create_col(ws_used, "phase")

# 기존 phase 값 비우기
if CLEAR_USED_DATA_PHASE:
    for r in range(2, ws_used.max_row + 1):
        ws_used.cell(row=r, column=used_phase_idx).value = None

# =========================
# 3-2) Copy non-native phase from gt_manual to existing used_data rows only
# =========================

USED_SHEET = "used_data"
CLEAR_USED_DATA_PHASE = True  # True면 used_data의 기존 phase 값을 먼저 비움

if USED_SHEET not in wb.sheetnames:
    raise ValueError(f"Sheet '{USED_SHEET}' not found. Existing sheets: {wb.sheetnames}")

ws_gt = wb[TARGET_SHEET]   # gt_manual
ws_used = wb[USED_SHEET]   # used_data


def get_headers(ws):
    return [cell.value for cell in ws[1]]


def get_col_idx(headers, candidates):
    col_name = find_col(headers, candidates)
    header_to_idx = {
        str(h).strip(): i + 1
        for i, h in enumerate(headers)
        if h is not None
    }
    return header_to_idx[str(col_name).strip()], col_name


def get_or_create_col(ws, col_name):
    headers = get_headers(ws)

    for i, h in enumerate(headers):
        if h is not None and norm_col(h) == norm_col(col_name):
            return i + 1

    new_idx = ws.max_column + 1
    ws.cell(row=1, column=new_idx).value = col_name
    return new_idx


# =========================
# used_data columns
# =========================
used_headers = get_headers(ws_used)

used_subject_idx, _ = get_col_idx(
    used_headers,
    ["anon_patient_ID", "subject_id", "subject", "patient"]
)
used_study_idx, _ = get_col_idx(
    used_headers,
    ["anon_study_ID", "study_id", "study"]
)
used_series_idx, _ = get_col_idx(
    used_headers,
    ["SeriesNumber", "series_id", "series"]
)

used_phase_idx = get_or_create_col(ws_used, "phase")


# =========================
# Build keys that already exist in used_data
# =========================
used_keys = set()

for r in range(2, ws_used.max_row + 1):
    subject = ws_used.cell(row=r, column=used_subject_idx).value
    study = ws_used.cell(row=r, column=used_study_idx).value
    series = ws_used.cell(row=r, column=used_series_idx).value

    key = make_key(subject, study, series)

    if key is not None:
        used_keys.add(key)

print(f"Existing used_data keys: {len(used_keys)}")


# =========================
# gt_manual columns
# =========================
gt_headers = get_headers(ws_gt)

gt_subject_idx, _ = get_col_idx(
    gt_headers,
    ["anon_patient_ID", "subject_id", "subject", "patient"]
)
gt_study_idx, _ = get_col_idx(
    gt_headers,
    ["anon_study_ID", "study_id", "study"]
)
gt_series_idx, _ = get_col_idx(
    gt_headers,
    ["SeriesNumber", "series_id", "series"]
)
gt_phase_idx, _ = get_col_idx(
    gt_headers,
    ["phase"]
)


# =========================
# Build non-native phase map ONLY for keys already in used_data
# =========================
non_native_phase_map = {}

for r in range(2, ws_gt.max_row + 1):
    subject = ws_gt.cell(row=r, column=gt_subject_idx).value
    study = ws_gt.cell(row=r, column=gt_study_idx).value
    series = ws_gt.cell(row=r, column=gt_series_idx).value
    phase_val = ws_gt.cell(row=r, column=gt_phase_idx).value

    key = make_key(subject, study, series)

    if key is None:
        continue

    # used_data에 없는 row는 무시
    if key not in used_keys:
        continue

    if phase_val is None:
        continue

    phase_str = str(phase_val).strip()

    if phase_str == "":
        continue

    # native는 복사하지 않음
    if phase_str.lower() == "native":
        continue

    non_native_phase_map[key] = phase_str

print(f"Non-native gt_manual phases matching existing used_data rows: {len(non_native_phase_map)}")


# =========================
# Clear existing used_data phase values if requested
# =========================
if CLEAR_USED_DATA_PHASE:
    for r in range(2, ws_used.max_row + 1):
        ws_used.cell(row=r, column=used_phase_idx).value = None

# =========================
# Copy phase only to existing used_data rows
# - non-native phase: copy actual phase
# - native phase: mark as "n/a"
# - missing/no match: leave blank
# =========================

# gt_manual에서 used_data에 이미 있는 key 중 native인 것 따로 저장
native_keys = set()

for r in range(2, ws_gt.max_row + 1):
    subject = ws_gt.cell(row=r, column=gt_subject_idx).value
    study = ws_gt.cell(row=r, column=gt_study_idx).value
    series = ws_gt.cell(row=r, column=gt_series_idx).value
    phase_val = ws_gt.cell(row=r, column=gt_phase_idx).value

    key = make_key(subject, study, series)

    if key is None:
        continue

    if key not in used_keys:
        continue

    if phase_val is None:
        continue

    phase_str = str(phase_val).strip()

    if phase_str.lower() == "native":
        native_keys.add(key)


# used_data에 phase 복사
copied_non_native = 0
marked_native_na = 0
missing_or_no_match = 0

for r in range(2, ws_used.max_row + 1):
    subject = ws_used.cell(row=r, column=used_subject_idx).value
    study = ws_used.cell(row=r, column=used_study_idx).value
    series = ws_used.cell(row=r, column=used_series_idx).value

    key = make_key(subject, study, series)

    if key is None:
        continue

    if key in non_native_phase_map:
        ws_used.cell(row=r, column=used_phase_idx).value = non_native_phase_map[key]
        copied_non_native += 1

    elif key in native_keys:
        ws_used.cell(row=r, column=used_phase_idx).value = "n/a"
        marked_native_na += 1

    else:
        ws_used.cell(row=r, column=used_phase_idx).value = None
        missing_or_no_match += 1

print(f"Copied non-native phases to existing used_data rows: {copied_non_native}")
print(f"Marked native rows as n/a in used_data: {marked_native_na}")
print(f"used_data rows left blank because missing/no match: {missing_or_no_match}")


# =========================
# Delete used_data rows where phase == "n/a"
# =========================

actual_native_count = marked_native_na
deleted_na_rows = 0

# row 삭제는 아래에서 위로 해야 index가 안 밀림
for r in range(ws_used.max_row, 1, -1):
    phase_val = ws_used.cell(row=r, column=used_phase_idx).value

    if phase_val is None:
        continue

    phase_str = str(phase_val).strip().lower()

    if phase_str in ["n/a", "na", "native"]:
        ws_used.delete_rows(r, 1)
        deleted_na_rows += 1

print("========================================")
print("Native deletion summary")
print("----------------------------------------")
print(f"{'Deleted n/a rows':<25} {'Actual native rows':<25}")
print(f"{deleted_na_rows:<25} {actual_native_count:<25}")
print("========================================")

# =========================
# 4) Save as new file
# =========================
wb.save(OUT_XLSX)
print(f"Saved: {OUT_XLSX}")